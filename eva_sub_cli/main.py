#!/usr/bin/env python
import csv
import json
import os
from collections import defaultdict

from ebi_eva_common_pyutils.config import WritableConfig
from openpyxl.reader.excel import load_workbook

from eva_sub_cli import SUB_CLI_CONFIG_FILE, __version__
from eva_sub_cli.validators.docker_validator import DockerValidator
from eva_sub_cli.validators.native_validator import NativeValidator
from eva_sub_cli.validators.validator import READY_FOR_SUBMISSION_TO_EVA
from eva_sub_cli.submit import StudySubmitter

VALIDATE = 'validate'
SUBMIT = 'submit'
DOCKER = 'docker'
NATIVE = 'native'


def get_vcf_files(mapping_file):
    vcf_files = []
    with open(mapping_file) as open_file:
        reader = csv.DictReader(open_file, delimiter=',')
        for row in reader:
            vcf_files.append(row['vcf'])
    return vcf_files


def create_vcf_files_mapping(submission_dir, vcf_files, reference_fasta, metadata_json, metadata_xlsx):
    mapping_file = os.path.join(submission_dir, 'vcf_mapping_file.csv')
    with open(mapping_file, 'w') as open_file:
        writer = csv.writer(open_file, delimiter=',')
        writer.writerow(['vcf', 'fasta', 'report'])
        if vcf_files and reference_fasta:
            for vcf_file in vcf_files:
                writer.writerow([os.path.abspath(vcf_file), os.path.abspath(reference_fasta)])
        elif metadata_json:
            create_vcf_files_mapping_from_metadata_json(writer, metadata_json)
        elif metadata_xlsx:
            create_vcf_files_mapping_from_metadata_xlsx(writer, metadata_xlsx)

    return mapping_file


def create_vcf_files_mapping_from_metadata_json(csv_writer, metadata_json):
    with open(metadata_json) as file:
        json_metadata = json.load(file)
    analysis_alias_dict = defaultdict(dict)
    for analysis in json_metadata['analysis']:
        analysis_alias_dict[analysis['analysisAlias']]['referenceFasta'] = analysis['referenceFasta']
        analysis_alias_dict[analysis['analysisAlias']]['assemblyReport'] = analysis['assemblyReport'] if 'assemblyReport' in analysis else ''

    for file in json_metadata['files']:
        reference_fasta = analysis_alias_dict[file['analysisAlias']]['referenceFasta']
        assembly_report = analysis_alias_dict[file['analysisAlias']]['assemblyReport']
        csv_writer.writerow([os.path.abspath(file['fileName']), os.path.abspath(reference_fasta), os.path.abspath(assembly_report) if assembly_report else ''])


def create_vcf_files_mapping_from_metadata_xlsx(csv_writer, metadata_xlsx):
    workbook = load_workbook(metadata_xlsx)

    analysis_alias_sheet = workbook['Analysis']
    analysis_headers = {}
    for cell in analysis_alias_sheet[1]:
        analysis_headers[cell.value] = cell.column - 1

    analysis_alias_dict = {}
    for row in analysis_alias_sheet.iter_rows(min_row=2, values_only=True):
        analysis_alias = row[analysis_headers['Analysis Alias']]
        reference_fasta = row[analysis_headers['Reference Fasta Path']]
        analysis_alias_dict[analysis_alias] = reference_fasta

    files_sheet = workbook['Files']
    files_headers = {}
    for cell in files_sheet[1]:
        files_headers[cell.value] = cell.column - 1

    for row in files_sheet.iter_rows(min_row=2, values_only=True):
        file_name = row[files_headers['File Name']]
        analysis_alias = row[files_headers['Analysis Alias']]
        reference_fasta = analysis_alias_dict[analysis_alias]
        csv_writer.writerow([os.path.abspath(file_name), os.path.abspath(reference_fasta)])


def orchestrate_process(submission_dir, vcf_files, reference_fasta, metadata_json, metadata_xlsx,
                        tasks, executor, username=None, password=None, **kwargs):
    # load config
    config_file_path = os.path.join(submission_dir, SUB_CLI_CONFIG_FILE)
    sub_config = WritableConfig(config_file_path, version=__version__)

    metadata_file = metadata_json or metadata_xlsx
    if not os.path.exists(os.path.abspath(metadata_file)):
        raise FileNotFoundError(f'The provided metadata file {metadata_file} does not exist')

    # Get the provided VCF and assembly
    vcf_files_mapping = create_vcf_files_mapping(submission_dir, vcf_files, reference_fasta, metadata_json, metadata_xlsx)
    vcf_files = get_vcf_files(vcf_files_mapping)

    # Validation is mandatory so if submit is requested then VALIDATE must have run before or be requested as well
    if SUBMIT in tasks and not sub_config.get(READY_FOR_SUBMISSION_TO_EVA, False):
        if VALIDATE not in tasks:
            tasks.append(VALIDATE)

    if VALIDATE in tasks:
        if executor == DOCKER:
            validator = DockerValidator(vcf_files_mapping, submission_dir, metadata_json, metadata_xlsx,
                                        submission_config=sub_config)
        # default to native execution
        else:
            validator = NativeValidator(vcf_files_mapping, submission_dir, metadata_json, metadata_xlsx,
                                        submission_config=sub_config)
        with validator:
            validator.validate_and_report()
            if not metadata_json:
                metadata_json = os.path.join(validator.output_dir, 'metadata.json')
            sub_config.set('metadata_json', value=metadata_json)
            sub_config.set('vcf_files', value=vcf_files)

    if SUBMIT in tasks:
        with StudySubmitter(submission_dir, submission_config=sub_config, username=username, password=password) as submitter:
            submitter.submit()
